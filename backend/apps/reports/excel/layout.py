"""
Layout-only module — pure rendering of one month sheet.

Inputs come from a fully-populated ``MonthData`` dataclass (see ``data.py``).
This module does NO DB access and NO data invention; that lives in ``data.py``.

The visual layout MUST stay byte-identical to the design samples produced by
``report_samples/generate_sample.py`` so that what the user sees in the
preview is what they actually download.
"""

from __future__ import annotations

import calendar
import datetime as dt
from dataclasses import dataclass, field
from decimal import Decimal
from typing import Optional

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.worksheet import Worksheet

# ──────────────────────────────────────────────────────────────────────────────
# CONSTANTS
# ──────────────────────────────────────────────────────────────────────────────

MONTH_NAMES_RU = [
    "Январь", "Февраль", "Март", "Апрель", "Май", "Июнь",
    "Июль", "Август", "Сентябрь", "Октябрь", "Ноябрь", "Декабрь",
]
WEEKDAY_RU = ["пн", "вт", "ср", "чт", "пт", "сб", "вс"]

# Colour palette
TEAL = "BDD7EE"
TEAL_DARK = "9DC3E6"
RED = "FF6E6E"
RED_LIGHT = "FCE4E4"
RED_DEEP = "C00000"
GREEN = "92D050"
GREEN_DARK = "548235"
BLUE_HDR = "4F81BD"
WHITE = "FFFFFF"
YELLOW = "FFE699"
BLACK = "000000"
ORANGE = "F8CBAD"

THIN = Side(border_style="thin", color="9C9C9C")
BORDERS = Border(top=THIN, bottom=THIN, left=THIN, right=THIN)

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
DATA = Alignment(horizontal="center", vertical="center", wrap_text=False)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=False)

NUMFMT = '#,##0;[Red]-#,##0;-'

# Column layout (no gutters)
DATE_COL = 1
DAY_ADMIN = 2
DAY_INCOME_COLS = list(range(3, 8))      # C–G  Общая сумма / Терминал / QR / Картага / Онлайн
DAY_EXPENSE_COLS = list(range(8, 14))    # H–M  6 expense categories
DAY_REMAIN_COL = 14
NIGHT_ADMIN = 15
NIGHT_INCOME_COLS = list(range(16, 21))
NIGHT_EXPENSE_COLS = list(range(21, 27))
NIGHT_REMAIN_COL = 27
PANEL_LBL_COL = 28
PANEL_FIRST_COL = 29
PANEL_COUNT = 3                          # always 3 admin columns

INCOME_HEADERS = ["Общая сумма", "Терминал", "QR Код", "Картага", "Онлайн"]
EXPENSE_HEADERS = [
    "Продукты", "Моющие средства", "Телеком-я харажати",
    "Ремонт", "Коммуналка", "Прочие",
]

# Payroll-row labels (Russian/Uzbek). Centralised so the same string
# isn't repeated across the panel-builder, formula resolver and totals row.
LBL_DAY = "День"
LBL_NIGHT = "Ночь"
LBL_TOTAL_SHIFT = "Общий смена"
LBL_FIX = "Фикса"
LBL_TOTAL_DAY = "Общий День"
LBL_TOTAL_NIGHT = "Общий Ночь"
LBL_BONUS = "Бонус"
LBL_FINE = "Жарима"
LBL_ADVANCE = "АВАНС"
LBL_REMAINING_SALARY = "Ойлик остатка"
LBL_TOTAL_SALARY = "Общий ойлик"
LBL_BONUS_PLUS = "Бонус +"
# Per-shift remainder header (one column at end of each shift block).
LBL_LEFTOVER = "Остатка"
# Lobar-only row inserted under the admin payroll panel for her director cut.
LBL_DIRECTOR_SALARY = "Директор маоши"

# (label, fill, font_color)
ADMIN_PAYROLL_ROWS: list[tuple[str, Optional[str], str]] = [
    (LBL_DAY,              TEAL,       BLACK),
    (LBL_NIGHT,            BLACK,      WHITE),
    (LBL_TOTAL_SHIFT,      GREEN,      BLACK),
    (LBL_FIX,              RED_LIGHT,  RED_DEEP),
    (LBL_TOTAL_DAY,        ORANGE,     BLACK),
    (LBL_TOTAL_NIGHT,      ORANGE,     BLACK),
    (LBL_BONUS,            RED_DEEP,   WHITE),
    (LBL_FINE,             RED,        BLACK),
    (LBL_ADVANCE,          BLUE_HDR,   WHITE),
    (LBL_REMAINING_SALARY, GREEN_DARK, WHITE),
    (LBL_TOTAL_SALARY,     None,       BLACK),
    (LBL_BONUS_PLUS,       YELLOW,     BLACK),
]

# ──────────────────────────────────────────────────────────────────────────────
# DATA CONTRACT — what the DB layer passes in
# ──────────────────────────────────────────────────────────────────────────────


@dataclass
class ShiftRow:
    """One half-day row (Day OR Night) for one calendar date."""
    admin_name: str = ""              # full name shown in the В/О column
    income_total: Decimal = Decimal(0)        # Общая сумма
    income_terminal: Decimal = Decimal(0)
    income_qr: Decimal = Decimal(0)
    income_card: Decimal = Decimal(0)         # Картага (card_transfer)
    income_online: Decimal = Decimal(0)
    expenses: list[Decimal] = field(default_factory=lambda: [Decimal(0)] * 6)
    # 6-item list aligned with EXPENSE_HEADERS


@dataclass
class StaffAttendanceRow:
    """One staff member's monthly attendance + payroll inputs."""
    full_name: str
    present_flags: list[bool]  # length = days_in_month
    fine: Decimal = Decimal(0)
    advance: Decimal = Decimal(0)


@dataclass
class AdminPanelInputs:
    """Per-admin manual numbers driving the right-hand payroll panel."""
    full_name: str
    fine: Decimal = Decimal(0)         # Жарима
    advance: Decimal = Decimal(0)      # АВАНС
    bonus_plus: Decimal = Decimal(0)   # Бонус +
    bonus_pct: float = 0.05            # 5 % default; 0 for general manager


@dataclass
class MonthData:
    year: int
    month: int
    branch_name: str
    lobar_variant: bool
    viewer: str                           # "CEO" or an admin's full_name
    daily_rate: int                       # staff per-shift rate (UZS)
    admin_shift_rate: int                 # admin Фикса rate per shift (UZS)
    director_salary: Decimal              # for Lobar's "Директор маоши" row
    day_rows: list[ShiftRow]              # length = days_in_month
    night_rows: list[ShiftRow]            # length = days_in_month
    admin_panels: list[AdminPanelInputs]  # length up to PANEL_COUNT
    staff: list[StaffAttendanceRow]


# ──────────────────────────────────────────────────────────────────────────────
# STYLE HELPERS
# ──────────────────────────────────────────────────────────────────────────────


def _fill(color: str) -> PatternFill:
    return PatternFill(start_color=color, end_color=color, fill_type="solid")


def _style(cell, *, bold=False, fill_color=None, font_color=None,
           number_format=None, align=DATA, border=BORDERS, size=10):
    cell.font = Font(bold=bold, color=font_color or BLACK,
                     name="Calibri", size=size)
    if fill_color:
        cell.fill = _fill(fill_color)
    if number_format:
        cell.number_format = number_format
    cell.alignment = align
    cell.border = border


def _col(c: int) -> str:
    return get_column_letter(c)


def _can_see(viewer: str, admin_name: str) -> bool:
    return bool(admin_name) and (viewer == "CEO" or admin_name == viewer)


# ──────────────────────────────────────────────────────────────────────────────
# RENDER ONE MONTH
# ──────────────────────────────────────────────────────────────────────────────


def build_month_sheet(ws: Worksheet, data: MonthData) -> None:  # noqa: C901,PLR0912,PLR0915
    year, month = data.year, data.month
    days = calendar.monthrange(year, month)[1]
    viewer = data.viewer
    lobar_variant = data.lobar_variant

    # Defensive: pad day/night rows to expected length
    day_rows = list(data.day_rows) + [ShiftRow() for _ in range(days - len(data.day_rows))]
    night_rows = list(data.night_rows) + [ShiftRow() for _ in range(days - len(data.night_rows))]

    ws.sheet_view.zoomScale = 100
    ws.sheet_view.zoomScaleNormal = 100

    # Column widths
    ws.column_dimensions[_col(DATE_COL)].width = 18
    ws.column_dimensions[_col(DAY_ADMIN)].width = 22
    for c in DAY_INCOME_COLS:
        ws.column_dimensions[_col(c)].width = 15
    for c in DAY_EXPENSE_COLS:
        ws.column_dimensions[_col(c)].width = 19
    ws.column_dimensions[_col(DAY_REMAIN_COL)].width = 15
    ws.column_dimensions[_col(NIGHT_ADMIN)].width = 22
    for c in NIGHT_INCOME_COLS:
        ws.column_dimensions[_col(c)].width = 15
    for c in NIGHT_EXPENSE_COLS:
        ws.column_dimensions[_col(c)].width = 19
    ws.column_dimensions[_col(NIGHT_REMAIN_COL)].width = 15
    ws.column_dimensions[_col(PANEL_LBL_COL)].width = 22
    for i in range(PANEL_COUNT):
        ws.column_dimensions[_col(PANEL_FIRST_COL + i)].width = 22

    # Row 1: super-headers
    ws.row_dimensions[1].height = 22
    ws.merge_cells(start_row=1, end_row=1,
                   start_column=DATE_COL, end_column=DAY_REMAIN_COL)
    _style(ws.cell(row=1, column=DATE_COL, value="ДЕНЬ"),
           bold=True, fill_color=TEAL, size=12)

    ws.merge_cells(start_row=1, end_row=1,
                   start_column=NIGHT_ADMIN, end_column=NIGHT_REMAIN_COL)
    _style(ws.cell(row=1, column=NIGHT_ADMIN, value="НОЧЬ"),
           bold=True, fill_color=BLACK, font_color=WHITE, size=12)

    panel_title = "ЛОБАР — Бош менеджер" if lobar_variant else "АДМИНЛАР"
    ws.merge_cells(start_row=1, end_row=1,
                   start_column=PANEL_LBL_COL,
                   end_column=PANEL_FIRST_COL + PANEL_COUNT - 1)
    _style(ws.cell(row=1, column=PANEL_LBL_COL, value=panel_title),
           bold=True, fill_color=BLUE_HDR, font_color=WHITE, size=12)

    # Row 2: column headers
    ws.row_dimensions[2].height = 42
    _style(ws.cell(row=2, column=DATE_COL,  value="Сана"),  bold=True, fill_color=TEAL)
    _style(ws.cell(row=2, column=DAY_ADMIN, value="Админ"), bold=True, fill_color=TEAL)
    for c, name in zip(DAY_INCOME_COLS, INCOME_HEADERS):
        _style(ws.cell(row=2, column=c, value=name), bold=True, fill_color=TEAL)
    for c, name in zip(DAY_EXPENSE_COLS, EXPENSE_HEADERS):
        _style(ws.cell(row=2, column=c, value=name),
               bold=True, fill_color=RED, font_color=WHITE)
    _style(ws.cell(row=2, column=DAY_REMAIN_COL, value=LBL_LEFTOVER),
           bold=True, fill_color=GREEN)

    _style(ws.cell(row=2, column=NIGHT_ADMIN, value="Админ"),
           bold=True, fill_color=TEAL)
    for c, name in zip(NIGHT_INCOME_COLS, INCOME_HEADERS):
        _style(ws.cell(row=2, column=c, value=name), bold=True, fill_color=TEAL)
    for c, name in zip(NIGHT_EXPENSE_COLS, EXPENSE_HEADERS):
        _style(ws.cell(row=2, column=c, value=name),
               bold=True, fill_color=RED, font_color=WHITE)
    _style(ws.cell(row=2, column=NIGHT_REMAIN_COL, value=LBL_LEFTOVER),
           bold=True, fill_color=GREEN)

    # Daily rows
    DATA_START = 3
    last_data_row = DATA_START + days - 1

    for i in range(days):
        row = DATA_START + i
        ws.row_dimensions[row].height = 18
        date_val = dt.date(year, month, i + 1)
        _style(ws.cell(row=row, column=DATE_COL, value=date_val),
               fill_color=TEAL, number_format="dd.mm.yyyy")

        # ── Day shift ─────────────────────────────────────────────────
        d = day_rows[i]
        _style(ws.cell(row=row, column=DAY_ADMIN, value=d.admin_name),
               fill_color=TEAL)
        ws.cell(row=row, column=DAY_INCOME_COLS[0], value=float(d.income_total))
        ws.cell(row=row, column=DAY_INCOME_COLS[1], value=float(d.income_terminal))
        ws.cell(row=row, column=DAY_INCOME_COLS[2], value=float(d.income_qr))
        ws.cell(row=row, column=DAY_INCOME_COLS[3], value=float(d.income_card))
        ws.cell(row=row, column=DAY_INCOME_COLS[4], value=float(d.income_online))
        for c, v in zip(DAY_EXPENSE_COLS, d.expenses):
            ws.cell(row=row, column=c, value=float(v))
        ws.cell(row=row, column=DAY_REMAIN_COL,
                value=(f"={_col(DAY_INCOME_COLS[0])}{row}"
                       f"-SUM({_col(DAY_EXPENSE_COLS[0])}{row}:"
                       f"{_col(DAY_EXPENSE_COLS[-1])}{row})"))
        for c in DAY_INCOME_COLS:
            _style(ws.cell(row=row, column=c), number_format=NUMFMT)
        for c in DAY_EXPENSE_COLS:
            _style(ws.cell(row=row, column=c),
                   fill_color=RED_LIGHT, number_format=NUMFMT)
        _style(ws.cell(row=row, column=DAY_REMAIN_COL),
               fill_color=GREEN, number_format=NUMFMT, bold=True)

        # ── Night shift ───────────────────────────────────────────────
        n = night_rows[i]
        _style(ws.cell(row=row, column=NIGHT_ADMIN, value=n.admin_name),
               fill_color=TEAL)
        ws.cell(row=row, column=NIGHT_INCOME_COLS[0], value=float(n.income_total))
        ws.cell(row=row, column=NIGHT_INCOME_COLS[1], value=float(n.income_terminal))
        ws.cell(row=row, column=NIGHT_INCOME_COLS[2], value=float(n.income_qr))
        ws.cell(row=row, column=NIGHT_INCOME_COLS[3], value=float(n.income_card))
        ws.cell(row=row, column=NIGHT_INCOME_COLS[4], value=float(n.income_online))
        for c, v in zip(NIGHT_EXPENSE_COLS, n.expenses):
            ws.cell(row=row, column=c, value=float(v))
        ws.cell(row=row, column=NIGHT_REMAIN_COL,
                value=(f"={_col(NIGHT_INCOME_COLS[0])}{row}"
                       f"-SUM({_col(NIGHT_EXPENSE_COLS[0])}{row}:"
                       f"{_col(NIGHT_EXPENSE_COLS[-1])}{row})"))
        for c in NIGHT_INCOME_COLS:
            _style(ws.cell(row=row, column=c), number_format=NUMFMT)
        for c in NIGHT_EXPENSE_COLS:
            _style(ws.cell(row=row, column=c),
                   fill_color=RED_LIGHT, number_format=NUMFMT)
        _style(ws.cell(row=row, column=NIGHT_REMAIN_COL),
               fill_color=GREEN, number_format=NUMFMT, bold=True)

    # Итог row
    total_row = last_data_row + 1
    _style(ws.cell(row=total_row, column=DATE_COL, value="Итог"),
           bold=True, fill_color=GREEN)
    _style(ws.cell(row=total_row, column=DAY_ADMIN), fill_color=GREEN)
    for c in DAY_INCOME_COLS + DAY_EXPENSE_COLS + [DAY_REMAIN_COL]:
        L = _col(c)
        ws.cell(row=total_row, column=c,
                value=f"=SUM({L}{DATA_START}:{L}{last_data_row})")
        _style(ws.cell(row=total_row, column=c),
               bold=True, fill_color=GREEN, number_format=NUMFMT)
    _style(ws.cell(row=total_row, column=NIGHT_ADMIN), fill_color=GREEN)
    for c in NIGHT_INCOME_COLS + NIGHT_EXPENSE_COLS + [NIGHT_REMAIN_COL]:
        L = _col(c)
        ws.cell(row=total_row, column=c,
                value=f"=SUM({L}{DATA_START}:{L}{last_data_row})")
        _style(ws.cell(row=total_row, column=c),
               bold=True, fill_color=GREEN, number_format=NUMFMT)

    # Recap mini-table
    recap_top = total_row + 1
    f_total_in_day = f"{_col(DAY_INCOME_COLS[0])}{total_row}"
    f_total_in_night = f"{_col(NIGHT_INCOME_COLS[0])}{total_row}"
    f_grand_in = f"({f_total_in_day}+{f_total_in_night})"
    f_exp_day = (f"SUM({_col(DAY_EXPENSE_COLS[0])}{total_row}:"
                 f"{_col(DAY_EXPENSE_COLS[-1])}{total_row})")
    f_exp_night = (f"SUM({_col(NIGHT_EXPENSE_COLS[0])}{total_row}:"
                   f"{_col(NIGHT_EXPENSE_COLS[-1])}{total_row})")
    f_grand_exp = f"({f_exp_day}+{f_exp_night})"
    cash_formula = (
        f"={f_grand_in}"
        f"-({_col(DAY_INCOME_COLS[1])}{total_row}+{_col(NIGHT_INCOME_COLS[1])}{total_row})"
        f"-({_col(DAY_INCOME_COLS[2])}{total_row}+{_col(NIGHT_INCOME_COLS[2])}{total_row})"
        f"-({_col(DAY_INCOME_COLS[3])}{total_row}+{_col(NIGHT_INCOME_COLS[3])}{total_row})"
        f"-({_col(DAY_INCOME_COLS[4])}{total_row}+{_col(NIGHT_INCOME_COLS[4])}{total_row})"
    )
    pct_formula = f"=ROUND({f_grand_exp}/{f_grand_in}*100,1)"
    pair1 = [
        ("Общий итог",     f"={f_grand_in}", GREEN, None),
        ("Общий расход",   f"={f_grand_exp}", GREEN, None),
        ("Продукта",       f"={_col(DAY_EXPENSE_COLS[0])}{total_row}+{_col(NIGHT_EXPENSE_COLS[0])}{total_row}", TEAL, None),
        ("Телеком харажи", f"={_col(DAY_EXPENSE_COLS[2])}{total_row}+{_col(NIGHT_EXPENSE_COLS[2])}{total_row}", TEAL, None),
        ("Коммуналка",     f"={_col(DAY_EXPENSE_COLS[4])}{total_row}+{_col(NIGHT_EXPENSE_COLS[4])}{total_row}", TEAL, None),
    ]
    pair2 = [
        (LBL_LEFTOVER, f"={_col(DAY_REMAIN_COL)}{total_row}+{_col(NIGHT_REMAIN_COL)}{total_row}", GREEN, None),
        ("%",        pct_formula, GREEN_DARK, "0.0"),
        ("Моющий",   f"={_col(DAY_EXPENSE_COLS[1])}{total_row}+{_col(NIGHT_EXPENSE_COLS[1])}{total_row}", TEAL, None),
        ("Ремонт",   f"={_col(DAY_EXPENSE_COLS[3])}{total_row}+{_col(NIGHT_EXPENSE_COLS[3])}{total_row}", TEAL, None),
        ("Прочие",   f"={_col(DAY_EXPENSE_COLS[5])}{total_row}+{_col(NIGHT_EXPENSE_COLS[5])}{total_row}", TEAL, None),
    ]
    pair3 = [
        ("QR Код",   f"={_col(DAY_INCOME_COLS[2])}{total_row}+{_col(NIGHT_INCOME_COLS[2])}{total_row}", TEAL, None),
        ("Картага",  f"={_col(DAY_INCOME_COLS[3])}{total_row}+{_col(NIGHT_INCOME_COLS[3])}{total_row}", TEAL, None),
        ("Накд",     cash_formula, TEAL, None),
        ("терминал", f"={_col(DAY_INCOME_COLS[1])}{total_row}+{_col(NIGHT_INCOME_COLS[1])}{total_row}", TEAL, None),
        ("Онлайн",   f"={_col(DAY_INCOME_COLS[4])}{total_row}+{_col(NIGHT_INCOME_COLS[4])}{total_row}", TEAL, None),
    ]
    pair_cols = [(1, 2), (3, 4), (5, 6)]
    for (lbl_col, val_col), pair in zip(pair_cols, [pair1, pair2, pair3]):
        for i, (label, formula, color, custom_fmt) in enumerate(pair):
            r = recap_top + i
            ws.cell(row=r, column=lbl_col, value=label)
            _style(ws.cell(row=r, column=lbl_col), bold=True, fill_color=color)
            ws.cell(row=r, column=val_col, value=formula)
            _style(ws.cell(row=r, column=val_col),
                   bold=True, number_format=(custom_fmt or NUMFMT),
                   fill_color="F2F2F2")

    # Admin payroll panel
    panels = list(data.admin_panels)[:PANEL_COUNT]
    while len(panels) < PANEL_COUNT:
        panels.append(AdminPanelInputs(full_name=""))
    admin_names = [p.full_name for p in panels]

    _style(ws.cell(row=2, column=PANEL_LBL_COL, value=""),
           fill_color=TEAL_DARK, align=CENTER)
    for i, name in enumerate(admin_names):
        c = PANEL_FIRST_COL + i
        ws.cell(row=2, column=c, value=name)
        _style(ws.cell(row=2, column=c),
               bold=True, fill_color=TEAL_DARK, align=CENTER)

    panel_row_for: dict[str, int] = {}
    for i, (label, fill_c, font_c) in enumerate(ADMIN_PAYROLL_ROWS):
        r = 3 + i
        panel_row_for[label] = r
        ws.cell(row=r, column=PANEL_LBL_COL, value=label)
        _style(ws.cell(row=r, column=PANEL_LBL_COL),
               bold=True, fill_color=fill_c, font_color=font_c, align=CENTER)
        for j, panel in enumerate(panels):
            c = PANEL_FIRST_COL + j
            name = panel.full_name
            if not name or not _can_see(viewer, name):
                ws.cell(row=r, column=c, value=None)
                _style(ws.cell(row=r, column=c), fill_color="D9D9D9")
                continue
            value = _panel_value(
                label=label, panel=panel,
                data_start=DATA_START, last_row=last_data_row,
                col_letter=_col(c), panel_row_for=panel_row_for,
                lobar_variant=lobar_variant,
                admin_shift_rate=data.admin_shift_rate,
            )
            ws.cell(row=r, column=c, value=value)
            _style(
                ws.cell(row=r, column=c),
                number_format=("0" if label in (LBL_DAY, LBL_NIGHT, LBL_TOTAL_SHIFT) else NUMFMT),
                bold=(label in (LBL_BONUS, LBL_REMAINING_SALARY, LBL_TOTAL_SALARY)),
            )

    extra_used = 0
    if lobar_variant:
        r = 3 + len(ADMIN_PAYROLL_ROWS)
        ws.cell(row=r, column=PANEL_LBL_COL, value=LBL_DIRECTOR_SALARY)
        _style(ws.cell(row=r, column=PANEL_LBL_COL),
               bold=True, fill_color=BLUE_HDR, font_color=WHITE)
        gm_name = panels[0].full_name
        if _can_see(viewer, gm_name):
            ws.cell(row=r, column=PANEL_FIRST_COL,
                    value=float(data.director_salary))
            _style(ws.cell(row=r, column=PANEL_FIRST_COL),
                   bold=True, number_format=NUMFMT)
        else:
            ws.cell(row=r, column=PANEL_FIRST_COL, value=None)
            _style(ws.cell(row=r, column=PANEL_FIRST_COL), fill_color="D9D9D9")
        for j in range(1, PANEL_COUNT):
            ws.cell(row=r, column=PANEL_FIRST_COL + j, value=None)
            _style(ws.cell(row=r, column=PANEL_FIRST_COL + j), fill_color="D9D9D9")
        panel_row_for[LBL_DIRECTOR_SALARY] = r
        extra_used = 1

    # Final black ИТОГ row
    final_row = 3 + len(ADMIN_PAYROLL_ROWS) + extra_used + 1
    ws.cell(row=final_row, column=PANEL_LBL_COL, value="ИТОГ")
    _style(ws.cell(row=final_row, column=PANEL_LBL_COL),
           bold=True, fill_color=BLACK, font_color=WHITE)
    for j, panel in enumerate(panels):
        c = PANEL_FIRST_COL + j
        L = _col(c)
        nm = panel.full_name
        if not nm or not _can_see(viewer, nm):
            ws.cell(row=final_row, column=c, value=None)
            _style(ws.cell(row=final_row, column=c),
                   bold=True, fill_color=BLACK, font_color=WHITE)
            continue
        oilik = panel_row_for[LBL_REMAINING_SALARY]
        bonusplus = panel_row_for[LBL_BONUS_PLUS]
        if lobar_variant and j == 0:
            director = panel_row_for[LBL_DIRECTOR_SALARY]
            formula = f"={L}{oilik}+{L}{bonusplus}+{L}{director}"
        else:
            formula = f"={L}{oilik}+{L}{bonusplus}"
        ws.cell(row=final_row, column=c, value=formula)
        _style(ws.cell(row=final_row, column=c),
               bold=True, fill_color=BLACK, font_color=WHITE,
               number_format=NUMFMT)

    # Staff attendance grid
    recap_bottom = recap_top + 4
    att_top = recap_bottom + 2
    _style(ws.cell(row=att_top, column=DATE_COL, value="Ходимлар"),
           bold=True, fill_color=BLUE_HDR, font_color=WHITE)
    for d in range(1, days + 1):
        c = DATE_COL + d
        ws.cell(row=att_top, column=c, value=f"{d:02d}.{month:02d}")
        _style(ws.cell(row=att_top, column=c),
               bold=True, fill_color=BLUE_HDR, font_color=WHITE)

    payroll_labels = [
        ("Умумий иш кунлари", TEAL,    BLACK),
        ("Жарима",            RED,     BLACK),
        ("АВАНС",             BLUE_HDR, WHITE),
        ("Общий ойлик",       GREEN,   BLACK),
        ("Колдик ойлик",      GREEN,   BLACK),
    ]
    payroll_first = DATE_COL + days + 1
    for i, (lbl, color, font_c) in enumerate(payroll_labels):
        c = payroll_first + i
        ws.merge_cells(start_row=att_top, end_row=att_top + 1,
                       start_column=c, end_column=c)
        ws.cell(row=att_top, column=c, value=lbl)
        _style(ws.cell(row=att_top, column=c),
               bold=True, fill_color=color, font_color=font_c)
        ws.column_dimensions[_col(c)].width = 20

    wd_row = att_top + 1
    _style(ws.cell(row=wd_row, column=DATE_COL), fill_color=BLUE_HDR)
    for d in range(1, days + 1):
        c = DATE_COL + d
        weekday = WEEKDAY_RU[dt.date(year, month, d).weekday()]
        ws.cell(row=wd_row, column=c, value=weekday)
        _style(ws.cell(row=wd_row, column=c),
               bold=True, fill_color=BLUE_HDR, font_color=WHITE)

    DAILY_RATE = data.daily_rate
    staff_first_row = wd_row + 1

    # Attendance dropdown ☑/☐
    if data.staff:
        tick_dv = DataValidation(type="list", formula1='"☑,☐"', allow_blank=True)
        tick_dv.error = "Фақат ☑ ёки ☐"
        tick_dv.errorTitle = "Нотўғри қиймат"
        tick_dv.prompt = "☑ = ишда,  ☐ = йўқ"
        tick_dv.promptTitle = "Давомат"
        first_att_letter = _col(DATE_COL + 1)
        last_att_letter = _col(DATE_COL + days)
        att_first_row = wd_row + 1
        att_last_row = att_first_row + len(data.staff) - 1
        tick_dv.add(f"{first_att_letter}{att_first_row}:{last_att_letter}{att_last_row}")
        ws.add_data_validation(tick_dv)

    for s_i, st in enumerate(data.staff):
        row = staff_first_row + s_i
        ws.cell(row=row, column=DATE_COL, value=st.full_name)
        _style(ws.cell(row=row, column=DATE_COL),
               bold=True, fill_color=TEAL, align=LEFT)
        flags = list(st.present_flags) + [False] * (days - len(st.present_flags))
        for d in range(1, days + 1):
            c = DATE_COL + d
            wd = dt.date(year, month, d).weekday()
            present = bool(flags[d - 1])
            ws.cell(row=row, column=c, value=("☑" if present else "☐"))
            cellfill = ("FFF2CC" if wd >= 5 else None)
            _style(ws.cell(row=row, column=c), fill_color=cellfill, size=14)

        first_letter = _col(DATE_COL + 1)
        last_letter = _col(DATE_COL + days)
        ws.cell(row=row, column=payroll_first,
                value=f'=COUNTIF({first_letter}{row}:{last_letter}{row},"☑")')
        _style(ws.cell(row=row, column=payroll_first), number_format="0")
        ws.cell(row=row, column=payroll_first + 1, value=float(st.fine))
        _style(ws.cell(row=row, column=payroll_first + 1), number_format=NUMFMT)
        ws.cell(row=row, column=payroll_first + 2, value=float(st.advance))
        _style(ws.cell(row=row, column=payroll_first + 2), number_format=NUMFMT)
        ws.cell(row=row, column=payroll_first + 3,
                value=f"={_col(payroll_first)}{row}*{DAILY_RATE}")
        _style(ws.cell(row=row, column=payroll_first + 3),
               number_format=NUMFMT, bold=True)
        ws.cell(row=row, column=payroll_first + 4,
                value=(f"={_col(payroll_first + 3)}{row}"
                       f"-{_col(payroll_first + 1)}{row}"
                       f"-{_col(payroll_first + 2)}{row}"))
        _style(ws.cell(row=row, column=payroll_first + 4),
               number_format=NUMFMT, bold=True, fill_color=GREEN)

    if data.staff:
        itog_row = staff_first_row + len(data.staff)
        ws.cell(row=itog_row, column=payroll_first - 1, value="ИТОГ")
        _style(ws.cell(row=itog_row, column=payroll_first - 1),
               bold=True, fill_color=GREEN)
        for i in range(5):
            c = payroll_first + i
            L = _col(c)
            ws.cell(row=itog_row, column=c,
                    value=f"=SUM({L}{staff_first_row}:{L}{itog_row - 1})")
            _style(ws.cell(row=itog_row, column=c),
                   bold=True, fill_color=GREEN, number_format=NUMFMT)

    ws.freeze_panes = "C3"


def _panel_value(*, label: str, panel: AdminPanelInputs,
                 data_start: int, last_row: int,
                 col_letter: str, panel_row_for: dict[str, int],
                 lobar_variant: bool, admin_shift_rate: int):
    name_quoted = f'"{panel.full_name}"'
    da = _col(DAY_ADMIN)
    na = _col(NIGHT_ADMIN)
    dt_ = _col(DAY_INCOME_COLS[0])
    nt_ = _col(NIGHT_INCOME_COLS[0])
    rng_day_admin = f"{da}{data_start}:{da}{last_row}"
    rng_night_admin = f"{na}{data_start}:{na}{last_row}"
    rng_day_total = f"{dt_}{data_start}:{dt_}{last_row}"
    rng_night_total = f"{nt_}{data_start}:{nt_}{last_row}"

    L = col_letter
    days_row = panel_row_for.get(LBL_DAY)
    nights_row = panel_row_for.get(LBL_NIGHT)
    fix_row = panel_row_for.get(LBL_FIX)
    obd_row = panel_row_for.get(LBL_TOTAL_DAY)
    obn_row = panel_row_for.get(LBL_TOTAL_NIGHT)
    bonus_row = panel_row_for.get(LBL_BONUS)
    fine_row = panel_row_for.get(LBL_FINE)
    adv_row = panel_row_for.get(LBL_ADVANCE)

    if label == LBL_DAY:
        return f"=COUNTA({rng_day_admin})" if lobar_variant else f'=COUNTIF({rng_day_admin},{name_quoted})'
    if label == LBL_NIGHT:
        return f"=COUNTA({rng_night_admin})" if lobar_variant else f'=COUNTIF({rng_night_admin},{name_quoted})'
    if label == LBL_TOTAL_SHIFT:
        return f"={L}{days_row}+{L}{nights_row}"
    if label == LBL_FIX:
        return f"={L}{panel_row_for[LBL_TOTAL_SHIFT]}*{admin_shift_rate}"
    if label == LBL_TOTAL_DAY:
        return f"=SUM({rng_day_total})" if lobar_variant else f'=SUMIF({rng_day_admin},{name_quoted},{rng_day_total})'
    if label == LBL_TOTAL_NIGHT:
        return f"=SUM({rng_night_total})" if lobar_variant else f'=SUMIF({rng_night_admin},{name_quoted},{rng_night_total})'
    if label == LBL_BONUS:
        pct = panel.bonus_pct
        return f"=ROUND(({L}{obd_row}+{L}{obn_row})*{pct},0)"
    if label == LBL_FINE:
        return float(panel.fine)
    if label == LBL_ADVANCE:
        return float(panel.advance)
    if label == LBL_REMAINING_SALARY:
        return (f"={L}{fix_row}+{L}{bonus_row}"
                f"-{L}{fine_row}-{L}{adv_row}")
    if label == LBL_TOTAL_SALARY:
        oilik_row = panel_row_for[LBL_REMAINING_SALARY]
        return f"={L}{adv_row}+{L}{oilik_row}"
    if label == LBL_BONUS_PLUS:
        return float(panel.bonus_plus)
    return 0
